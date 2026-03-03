{\rtf1\ansi\ansicpg1252\cocoartf2867
\cocoatextscaling0\cocoaplatform0{\fonttbl\f0\fswiss\fcharset0 Helvetica;\f1\fnil\fcharset0 LucidaGrande;}
{\colortbl;\red255\green255\blue255;}
{\*\expandedcolortbl;;}
\margl1440\margr1440\vieww11520\viewh8400\viewkind0
\pard\tx720\tx1440\tx2160\tx2880\tx3600\tx4320\tx5040\tx5760\tx6480\tx7200\tx7920\tx8640\pardirnatural\partightenfactor0

\f0\fs24 \cf0 #!/usr/bin/env python3\
"""\
NYC Multifamily Property Manager Intelligence & Targeting System\
LIVE DATA PIPELINE \'97 v3.0 (local-CSV only, no network calls)\
\
Usage\
\uc0\u9472 \u9472 \u9472 \u9472 \u9472 \
    python3 nyc_live_pipeline.py --data-dir /path/to/csvs [--output-dir .]\
\
Required files in --data-dir\
\uc0\u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \
    hpd_contacts.csv        
\f1 \uc0\u8592 
\f0  feu5-w2e2  (HPD Registration Contacts)\
    hpd_registrations.csv   
\f1 \uc0\u8592 
\f0  tesw-yqqr  (HPD Multiple Dwelling Registrations)\
    pluto.csv               
\f1 \uc0\u8592 
\f0  64uk-42ks  (MapPLUTO latest)\
\
Optional (Phase 2 violations scoring):\
    hpd_violations.csv      
\f1 \uc0\u8592 
\f0  wvxf-dwi5  (HPD Violations)\
\
Download commands (run outside this sandbox):\
\uc0\u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \
    export TOK="<your NYC Open Data App Token>"\
    BASE="https://data.cityofnewyork.us/resource"\
    curl -L -H "X-App-Token:$TOK" -o hpd_contacts.csv      "$BASE/feu5-w2e2.csv?\\$limit=750000"\
    curl -L -H "X-App-Token:$TOK" -o hpd_registrations.csv "$BASE/tesw-yqqr.csv?\\$limit=250000"\
    curl -L -H "X-App-Token:$TOK" -o pluto.csv             "$BASE/64uk-42ks.csv?\\$limit=1000000"\
    curl -L -H "X-App-Token:$TOK" -o hpd_violations.csv    "$BASE/wvxf-dwi5.csv?\\$limit=1200000"\
\
Outputs\
\uc0\u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \
    NYC_Multifamily_TargetManagers_Phase1.xlsx\
        Sheets: Ranked_Managers | Raw_Buildings | Data_Notes | Run_Log\
\
    NYC_Multifamily_TargetManagers_Phase2_Scored.xlsx\
        Sheets: Ranked_Managers_Scored | Raw_Buildings_Scored | Data_Notes | Run_Log\
"""\
\
import argparse\
import hashlib\
import os\
import re\
import sys\
from datetime import datetime\
from pathlib import Path\
\
import numpy as np\
import pandas as pd\
from openpyxl import Workbook\
from openpyxl.styles import Alignment, Font, PatternFill\
from openpyxl.utils import get_column_letter\
\
import warnings\
warnings.filterwarnings("ignore")\
\
# \uc0\u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \
# CONFIGURATION\
# \uc0\u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \
\
TARGET_ZIPS = \{"10021", "10028", "10075", "11209", "11375", "11374"\}\
\
ZIP_TO_NEIGHBORHOOD = \{\
    "10021": "Upper East Side (Manhattan)",\
    "10028": "Upper East Side (Manhattan)",\
    "10075": "Upper East Side (Manhattan)",\
    "11209": "Bay Ridge (Brooklyn)",\
    "11375": "Forest Hills / Rego Park (Queens)",\
    "11374": "Forest Hills / Rego Park (Queens)",\
\}\
\
MIN_UNITS = 25\
MAX_UNITS = 120\
\
SOQL_URLS = \{\
    "hpd_contacts.csv":      "https://data.cityofnewyork.us/resource/feu5-w2e2.csv?$limit=750000",\
    "hpd_registrations.csv": "https://data.cityofnewyork.us/resource/tesw-yqqr.csv?$limit=250000",\
    "pluto.csv":             "https://data.cityofnewyork.us/resource/64uk-42ks.csv?$limit=1000000",\
    "hpd_violations.csv":    "https://data.cityofnewyork.us/resource/wvxf-dwi5.csv?$limit=1200000",\
\}\
\
# \uc0\u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \
# RUN-LOG\
# \uc0\u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \
\
_run_log: list[dict] = []\
\
def log(step: str, value, note: str = "") -> None:\
    ts = datetime.now().strftime("%H:%M:%S")\
    _run_log.append(\{"Timestamp": ts, "Step": step, "Value": str(value), "Notes": note\})\
    print(f"  [\{ts\}] \{step:<50\} \{value\}" + (f"  \'97 \{note\}" if note else ""))\
\
\
# \uc0\u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \
# FATAL HELPER  (no network, no synthetic data \'97 just die clearly)\
# \uc0\u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \
\
def fatal(msg: str) -> None:\
    border = "=" * 70\
    print(f"\\n\{border\}\\nFATAL: \{msg\}\\n\{border\}\\n", file=sys.stderr)\
    sys.exit(1)\
\
\
# \uc0\u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \
# LOCAL CSV LOADER  (hard-fails on missing file; no download attempted)\
# \uc0\u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \
\
def sha256(path: str) -> str:\
    h = hashlib.sha256()\
    with open(path, "rb") as f:\
        for chunk in iter(lambda: f.read(1 << 20), b""):\
            h.update(chunk)\
    return h.hexdigest()\
\
\
def load_csv(label: str, path: str, required: bool = True) -> pd.DataFrame | None:\
    """\
    Load a local CSV into a DataFrame.\
    If required=True and file is absent 
\f1 \uc0\u8594 
\f0  FATAL (no network fallback, ever).\
    If required=False and absent 
\f1 \uc0\u8594 
\f0  returns None and logs a warning.\
    """\
    p = Path(path)\
    if not p.exists():\
        if required:\
            fatal(\
                f"\{label\} not found: \{path\}\\n\\n"\
                f"  This pipeline never makes network calls. Supply the file at the\\n"\
                f"  path above, or point --data-dir to the directory containing it.\\n\\n"\
                f"  Download command (run outside this sandbox):\\n"\
                f"    curl -L -H 'X-App-Token:<TOKEN>' \\\\\\n"\
                f"         -o '\{p.name\}' \\\\\\n"\
                f"         '\{SOQL_URLS.get(p.name, '<see script header>')\}'"\
            )\
        log(f"\{label\} (SKIPPED \'97 file absent)", "n/a", path)\
        return None\
\
    size_mb = p.stat().st_size / (1 << 20)\
    chk = sha256(path)\
    df = pd.read_csv(path, dtype=str, low_memory=False)\
    log(f"Loaded \{label\}", f"\{len(df):,\} rows", f"\{size_mb:.1f\} MB  sha256=\{chk[:16]\}\'85")\
    return df\
\
\
# \uc0\u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \
# COLUMN FINDER\
# \uc0\u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \
\
def col(df: pd.DataFrame, *candidates: str) -> str | None:\
    """Return first matching column name (case-insensitive), or None."""\
    lc = \{c.lower(): c for c in df.columns\}\
    for c in candidates:\
        if c.lower() in lc:\
            return lc[c.lower()]\
    return None\
\
\
def require_col(df: pd.DataFrame, label: str, *candidates: str) -> str:\
    c = col(df, *candidates)\
    if c is None:\
        fatal(\
            f"Cannot find column '\{candidates[0]\}' (or synonyms) in \{label\}.\\n"\
            f"  Columns present: \{list(df.columns)\}"\
        )\
    return c\
\
\
# \uc0\u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \
# BBL NORMALIZATION\
# \uc0\u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \
\
BORO_ALPHA = \{"MN": "1", "BX": "2", "BK": "3", "QN": "4", "SI": "5"\}\
BORO_NAME  = \{"MANHATTAN": "1", "BRONX": "2", "BROOKLYN": "3",\
              "QUEENS": "4", "STATEN ISLAND": "5"\}\
\
\
def _clean_int(v, w: int) -> str | None:\
    try:\
        return str(int(float(str(v).strip()))).zfill(w)\
    except (ValueError, TypeError):\
        return None\
\
\
def normalize_bbl(raw) -> str | None:\
    if pd.isna(raw):\
        return None\
    digits = re.sub(r"\\D", "", str(raw))\
    return digits.zfill(10) if len(digits) >= 6 else None\
\
\
def construct_bbl(boro, block, lot) -> str | None:\
    def resolve_boro(v):\
        v = str(v).strip().upper()\
        if v.isdigit():\
            return v.zfill(1)\
        return BORO_ALPHA.get(v) or BORO_NAME.get(v)\
\
    b  = resolve_boro(boro)\
    bl = _clean_int(block, 5)\
    l  = _clean_int(lot, 4)\
    return (b + bl + l) if (b and bl and l) else None\
\
\
def add_bbl(df: pd.DataFrame, label: str) -> pd.DataFrame:\
    """Add bbl_norm column; prefers existing bbl field, falls back to boro+block+lot."""\
    df = df.copy()\
\
    bbl_c = col(df, "bbl")\
    if bbl_c:\
        df["bbl_norm"] = df[bbl_c].apply(normalize_bbl)\
        ok = df["bbl_norm"].notna().sum()\
        log(f"BBL (\{label\})", f"\{ok:,\}/\{len(df):,\} valid", f"from existing '\{bbl_c\}' column")\
        if ok / max(len(df), 1) > 0.5:\
            return df\
\
    boro_c  = col(df, "boroid", "boro", "borough", "borocode")\
    block_c = col(df, "block")\
    lot_c   = col(df, "lot")\
    if boro_c and block_c and lot_c:\
        df["bbl_norm"] = [\
            construct_bbl(b, bl, l)\
            for b, bl, l in zip(df[boro_c], df[block_c], df[lot_c])\
        ]\
        ok = df["bbl_norm"].notna().sum()\
        log(f"BBL (\{label\})", f"\{ok:,\}/\{len(df):,\} valid",\
            f"constructed from \{boro_c\}+\{block_c\}+\{lot_c\}")\
    else:\
        df["bbl_norm"] = None\
        log(f"BBL (\{label\})", "NONE", "boro/block/lot columns not found")\
\
    return df\
\
\
# \uc0\u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \
# ADDRESS NORMALIZER (for fallback join)\
# \uc0\u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \
\
def norm_addr(s) -> str:\
    if pd.isna(s):\
        return ""\
    s = re.sub(r"\\bSTREET\\b", "ST", str(s).upper())\
    s = re.sub(r"\\bAVENUE\\b", "AVE", s)\
    s = re.sub(r"\\bBOULEVARD\\b", "BLVD", s)\
    s = re.sub(r"[^\\w\\s]", " ", s)\
    return " ".join(s.split()).strip()\
\
\
# \uc0\u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \
# AGENT NAME NORMALIZER (deduplication key only)\
# \uc0\u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \
\
def norm_name(name) -> str:\
    if pd.isna(name) or not str(name).strip():\
        return "UNKNOWN"\
    n = re.sub(\
        r"\\b(LLC|INC|CORP|LTD|CO\\.?|COMPANY|ASSOCIATES|ASSOC\\.?"\
        r"|MGMT|MANAGEMENT|REALTY|PROPERTIES|GROUP|SERVICES|SVCS)\\b",\
        "", str(name).upper())\
    return " ".join(re.sub(r"[^\\w\\s]", " ", n).split())\
\
\
# \uc0\u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \
# EXCEL HELPERS\
# \uc0\u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \
\
def _auto_width(ws, min_w=8, max_w=58):\
    for col_cells in ws.columns:\
        best = min_w\
        for cell in col_cells:\
            try:\
                best = max(best, min(len(str(cell.value or "")) + 2, max_w))\
            except Exception:\
                pass\
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = best\
\
\
def _write_sheet(ws, df: pd.DataFrame,\
                 hdr_color: str = "1F4E79",\
                 note: str = "") -> None:\
    """Write note row (1), header row (2), data rows (3+)."""\
    n_cols = len(df.columns)\
\
    if note:\
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)\
        c = ws["A1"]\
        c.value = note\
        c.font = Font(bold=True, color="7B1FA2", size=9)\
        c.fill = PatternFill("solid", fgColor="F3E5F5")\
        c.alignment = Alignment(wrap_text=True)\
        ws.row_dimensions[1].height = 38\
        hdr_row = 2\
    else:\
        hdr_row = 1\
\
    fill_hdr = PatternFill("solid", fgColor=hdr_color)\
    fill_alt = PatternFill("solid", fgColor="EBF3FB")\
    for ci, name in enumerate(df.columns, 1):\
        c = ws.cell(row=hdr_row, column=ci, value=name)\
        c.fill = fill_hdr\
        c.font = Font(bold=True, color="FFFFFF")\
        c.alignment = Alignment(horizontal="center", wrap_text=True)\
\
    data_start = hdr_row + 1\
    for ri, row_vals in enumerate(df.itertuples(index=False)):\
        r = data_start + ri\
        for ci, v in enumerate(row_vals, 1):\
            ws.cell(row=r, column=ci, value=v)\
        if ri % 2 == 0:\
            for ci in range(1, n_cols + 1):\
                ws.cell(row=r, column=ci).fill = fill_alt\
\
    ws.freeze_panes = ws.cell(row=data_start, column=1)\
    _auto_width(ws)\
\
\
TIER_CLR = \{"A": "C8E6C9", "B": "FFF9C4", "C": "FFCCBC"\}\
\
\
# \uc0\u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \
# PIPELINE\
# \uc0\u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \
\
def run(data_dir: str, output_dir: str) -> None:\
    os.makedirs(output_dir, exist_ok=True)\
    run_start = datetime.now().isoformat()\
    log("Pipeline start", run_start)\
\
    # \uc0\u9472 \u9472  LOAD \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \
    print("\\n=== LOAD ===")\
    contacts = load_csv("HPD Contacts",      os.path.join(data_dir, "hpd_contacts.csv"),      required=True)\
    regs     = load_csv("HPD Registrations", os.path.join(data_dir, "hpd_registrations.csv"), required=True)\
    pluto    = load_csv("PLUTO",             os.path.join(data_dir, "pluto.csv"),             required=True)\
    viols    = load_csv("HPD Violations",    os.path.join(data_dir, "hpd_violations.csv"),    required=False)\
\
    # \uc0\u9472 \u9472  INSPECT COLUMNS \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \
    print("\\n=== COLUMN MAP ===")\
    for label, df in [("HPD Contacts", contacts), ("HPD Regs", regs), ("PLUTO", pluto)]:\
        print(f"  \{label\}: \{list(df.columns)\}")\
\
    # \uc0\u9472 \u9472  BUILD BBLs \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \
    print("\\n=== BBL NORMALIZATION ===")\
    regs  = add_bbl(regs,  "HPD Regs")\
    pluto = add_bbl(pluto, "PLUTO")\
\
    # \uc0\u9472 \u9472  IDENTIFY CRITICAL COLUMNS \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \
    print("\\n=== COLUMN IDENTIFICATION ===")\
    # Contacts\
    c_regid  = require_col(contacts, "HPD Contacts", "registrationid")\
    c_type   = require_col(contacts, "HPD Contacts", "type")\
    c_corp   = col(contacts, "corporationname")\
    c_fname  = col(contacts, "firstname")\
    c_lname  = col(contacts, "lastname")\
    c_bhouse = col(contacts, "businesshousenumber")\
    c_bstreet= col(contacts, "businessstreetname")\
    c_bapt   = col(contacts, "businessapartment")\
    c_bcity  = col(contacts, "businesscity")\
    c_bstate = col(contacts, "businessstate")\
    c_bzip   = col(contacts, "businesszip")\
    c_bphone = col(contacts, "businessphone")\
\
    # Registrations\
    r_regid  = require_col(regs, "HPD Regs", "registrationid")\
    r_zip    = require_col(regs, "HPD Regs", "zip", "zipcode")\
    r_lc     = col(regs, "lifecycle")\
    r_house  = col(regs, "housenumber", "buildinghousenumber")\
    r_street = col(regs, "streetname",  "buildingstreetname")\
    r_boro   = col(regs, "boro", "borough", "boroid")\
\
    # PLUTO\
    p_zip    = require_col(pluto, "PLUTO", "postcode", "zipcode", "zip")\
    p_units  = require_col(pluto, "PLUTO", "unitsres")\
    p_addr   = col(pluto, "address")\
    p_lu     = col(pluto, "landuse")\
    p_bc     = col(pluto, "bldgclass")\
    p_yb     = col(pluto, "yearbuilt")\
    p_boro   = col(pluto, "borough", "boro")\
    p_retail = col(pluto, "retailarea")\
\
    for label, val in [\
        ("contacts.registrationid", c_regid), ("contacts.type", c_type),\
        ("regs.registrationid", r_regid),     ("regs.zip", r_zip),\
        ("regs.lifecycle", r_lc),\
        ("pluto.postcode/zipcode", p_zip),    ("pluto.unitsres", p_units),\
        ("pluto.landuse", p_lu),\
    ]:\
        print(f"    \{label:<35\} 
\f1 \uc0\u8594 
\f0  \{val\}")\
\
    # \uc0\u9472 \u9472  FILTER CONTACTS TO AGENT TYPE \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \
    print("\\n=== FILTER CONTACTS (type=Agent) ===")\
    log("Contacts raw", len(contacts))\
    agents = contacts[contacts[c_type].str.strip().str.lower() == "agent"].copy()\
    log("Contacts type=Agent", len(agents))\
\
    def display_name(row):\
        corp = str(row.get(c_corp, "") or "").strip() if c_corp else ""\
        if corp and corp.upper() not in ("", "NAN", "NONE"):\
            return corp\
        fn = str(row.get(c_fname, "") or "").strip() if c_fname else ""\
        ln = str(row.get(c_lname, "") or "").strip() if c_lname else ""\
        return f"\{fn\} \{ln\}".strip() or "UNKNOWN"\
\
    def mailing(row):\
        parts = [\
            str(row.get(f, "") or "").strip()\
            for f in [c_bhouse, c_bstreet, c_bapt, c_bcity, c_bstate, c_bzip]\
            if f\
        ]\
        return " ".join(p for p in parts if p and p.upper() != "NAN")\
\
    agents["AgentName"]     = agents.apply(display_name, axis=1)\
    agents["MailingAddr"]   = agents.apply(mailing, axis=1)\
    agents["Phone"]         = agents[c_bphone].fillna("") if c_bphone else ""\
    agents[c_regid]         = agents[c_regid].astype(str).str.strip()\
\
    # \uc0\u9472 \u9472  FILTER REGS TO ACTIVE + TARGET ZIPs \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \
    print("\\n=== FILTER REGISTRATIONS ===")\
    log("Regs raw", len(regs))\
    regs[r_zip] = regs[r_zip].astype(str).str.strip().str.zfill(5)\
\
    if r_lc:\
        active = regs[regs[r_lc].str.strip().str.lower() == "active"].copy()\
        log("Regs Active", len(active))\
    else:\
        active = regs.copy()\
        log("Regs (no lifecycle column \'97 using all)", len(active))\
\
    active_tgt = active[active[r_zip].isin(TARGET_ZIPS)].copy()\
    log("Regs Active + target ZIPs", len(active_tgt))\
    active_tgt[r_regid] = active_tgt[r_regid].astype(str).str.strip()\
\
    # \uc0\u9472 \u9472  JOIN CONTACTS 
\f1 \uc0\u8594 
\f0  REGISTRATIONS \uc0\u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \
    print("\\n=== JOIN: Contacts \'d7 Regs (on registrationid) ===")\
    reg_pull = [r_regid, "bbl_norm"]\
    for c2 in [r_zip, r_house, r_street, r_boro]:\
        if c2 and c2 not in reg_pull:\
            reg_pull.append(c2)\
    reg_pull = list(dict.fromkeys(reg_pull))\
\
    cr = agents.merge(\
        active_tgt[reg_pull],\
        left_on=c_regid, right_on=r_regid,\
        how="inner", suffixes=("", "_reg")\
    )\
    log("Contacts \'d7 Regs inner join", len(cr))\
\
    # Assign neighborhood from building ZIP\
    zip_src = r_zip if r_zip in cr.columns else None\
    if zip_src:\
        cr["Neighborhood"] = cr[zip_src].map(ZIP_TO_NEIGHBORHOOD)\
        cr = cr[cr["Neighborhood"].notna()].copy()\
        log("After neighborhood ZIP filter", len(cr))\
\
    # \uc0\u9472 \u9472  FILTER PLUTO TO TARGET ZIPs \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \
    print("\\n=== FILTER PLUTO ===")\
    pluto[p_zip] = pluto[p_zip].astype(str).str.strip().str.zfill(5)\
    pluto_tgt = pluto[pluto[p_zip].isin(TARGET_ZIPS)].copy()\
    log("PLUTO raw", len(pluto))\
    log("PLUTO target ZIPs", len(pluto_tgt))\
    pluto_tgt[p_units] = pd.to_numeric(pluto_tgt[p_units], errors="coerce")\
\
    # \uc0\u9472 \u9472  JOIN HPD+CONTACTS 
\f1 \uc0\u8594 
\f0  PLUTO (BBL primary; address+ZIP fallback) \uc0\u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \
    print("\\n=== JOIN: HPD \'d7 PLUTO ===")\
    pluto_cols = list(dict.fromkeys(\
        c2 for c2 in ["bbl_norm", p_zip, p_units, p_addr,\
                       p_lu, p_bc, p_yb, p_boro, p_retail]\
        if c2 is not None\
    ))\
    pluto_slim = pluto_tgt[pluto_cols].copy()\
    if p_addr:\
        pluto_slim["_addr_norm"] = pluto_slim[p_addr].apply(norm_addr)\
\
    # Primary: BBL join\
    merged = cr.merge(pluto_slim, on="bbl_norm", how="inner", suffixes=("", "_pluto"))\
    n_bbl = len(merged)\
    log("BBL join rows", n_bbl)\
\
    # Fallback: address+ZIP for unmatched BBLs\
    unmatched = cr[~cr["bbl_norm"].isin(merged["bbl_norm"].dropna()) |\
                    cr["bbl_norm"].isna()].copy()\
    n_fallback = 0\
\
    if len(unmatched) and r_house and r_street and p_addr:\
        unmatched["_addr_norm"] = (\
            unmatched[r_house].fillna("") + " " +\
            unmatched[r_street].fillna("")\
        ).apply(norm_addr)\
        join_keys = ["_addr_norm"]\
        if p_zip in unmatched.columns and p_zip in pluto_slim.columns:\
            join_keys.append(p_zip)\
        fb = unmatched.merge(\
            pluto_slim, on=join_keys, how="inner", suffixes=("", "_pluto")\
        )\
        n_fallback = len(fb)\
        log("Address+ZIP fallback rows", n_fallback)\
        merged = pd.concat([merged, fb], ignore_index=True)\
        if "bbl_norm" in merged.columns:\
            merged = merged.drop_duplicates(subset=["bbl_norm"])\
    else:\
        log("Address+ZIP fallback", 0, "skipped (insufficient columns)")\
\
    log("Total after join", len(merged))\
\
    # \uc0\u9472 \u9472  UNIT FILTER \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \
    print("\\n=== UNIT FILTER ===")\
    merged[p_units] = pd.to_numeric(merged[p_units], errors="coerce")\
    filtered = merged[merged[p_units].between(MIN_UNITS, MAX_UNITS)].copy()\
    log("After UnitsRes 25\'96120 filter", len(filtered))\
\
    if len(filtered) == 0:\
        fatal(\
            "Zero buildings remain after filters.\\n"\
            f"  Rows before filter: \{len(merged)\}\\n"\
            f"  Check: (a) join quality \'97 BBL format match between HPD and PLUTO?\\n"\
            f"         (b) target ZIPs present in both datasets?\\n"\
            f"         (c) UnitsRes column populated in PLUTO file?"\
        )\
\
    # \uc0\u9472 \u9472  NEIGHBORHOOD \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \
    if "Neighborhood" not in filtered.columns:\
        zip_final = p_zip if p_zip in filtered.columns else zip_src\
        if zip_final:\
            filtered["Neighborhood"] = filtered[zip_final].map(ZIP_TO_NEIGHBORHOOD)\
\
    # \uc0\u9472 \u9472  BUILDING ADDRESS \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \
    if p_addr and p_addr in filtered.columns:\
        filtered["BuildingAddr"] = filtered[p_addr].fillna("")\
    elif r_house and r_street:\
        filtered["BuildingAddr"] = (\
            filtered[r_house].fillna("") + " " + filtered[r_street].fillna("")\
        ).str.strip()\
    else:\
        filtered["BuildingAddr"] = filtered.get("bbl_norm", "")\
\
    zip_col  = p_zip   if p_zip  in filtered.columns else (zip_src or "")\
    boro_col = p_boro  if (p_boro and p_boro in filtered.columns) else (r_boro or "")\
\
    # \uc0\u9472 \u9472  AGGREGATE BY AGENT \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \
    print("\\n=== AGGREGATE BY AGENT ===")\
    filtered["_norm"] = filtered["AgentName"].apply(norm_name)\
\
    # Density: 3+ buildings in same ZIP\
    dense_set = set(\
        filtered.groupby(["_norm", zip_col]).size()\
        .reset_index(name="n").query("n >= 3")["_norm"]\
    ) if zip_col else set()\
\
    def join_uniq(s):\
        return " | ".join(sorted(\{str(v) for v in s if pd.notna(v)\}))\
\
    def join_zips(s):\
        return ", ".join(sorted(\{str(v) for v in s if pd.notna(v)\}))\
\
    grp = filtered.groupby("_norm").agg(\
        AgentName   =("AgentName",   "first"),\
        Mailing     =("MailingAddr", "first"),\
        Phone       =("Phone",       "first"),\
        Neighborhoods=("Neighborhood", join_uniq),\
        ZIPs        =(zip_col,       join_zips) if zip_col else ("AgentName", lambda _: ""),\
        N_Bldgs     =(p_units,       "count"),\
        TotalUnits  =(p_units,       "sum"),\
        AvgUnits    =(p_units,       "mean"),\
    ).reset_index()\
\
    grp["PortfolioFlag"] = (grp["N_Bldgs"] >= 5).map(\{True: "Yes", False: "No"\})\
    grp["MultiZIPFlag"]  = grp["ZIPs"].apply(\
        lambda z: "Yes" if len(z.split(",")) >= 2 else "No")\
    grp["DensityFlag"]   = grp["_norm"].apply(\
        lambda n: "Yes" if n in dense_set else "No")\
    grp["Notes"] = grp["DensityFlag"].apply(\
        lambda f: "Density: same-ZIP heuristic (\uc0\u8805 3 bldgs/ZIP); no geocoding." if f == "Yes" else "")\
\
    ranked_p1 = pd.DataFrame(\{\
        "Managing Agent Name":          grp["AgentName"],\
        "Mailing Address":              grp["Mailing"],\
        "Phone":                        grp["Phone"],\
        "Neighborhood":                 grp["Neighborhoods"],\
        "ZIP Code(s)":                  grp["ZIPs"],\
        "# Buildings Managed":          grp["N_Bldgs"],\
        "Total Units":                  grp["TotalUnits"].astype(int),\
        "Avg Units per Building":       grp["AvgUnits"].round(1),\
        "Portfolio Operator Flag (5+)": grp["PortfolioFlag"],\
        "Multi-ZIP Flag":               grp["MultiZIPFlag"],\
        "Density Flag":                 grp["DensityFlag"],\
        "Notes":                        grp["Notes"],\
    \}).drop_duplicates(subset=["Managing Agent Name"])\
\
    ranked_p1 = ranked_p1.sort_values(\
        ["# Buildings Managed", "Total Units"], ascending=[False, False]\
    ).reset_index(drop=True)\
    log("Phase 1 ranked managers", len(ranked_p1))\
\
    # Raw buildings sheet\
    raw_map = \{\
        "BuildingAddr": "Address",\
        "bbl_norm":     "BBL (if available)",\
        boro_col:       "Borough",\
        zip_col:        "ZIP",\
        p_units:        "UnitsRes",\
        "AgentName":    "Managing Agent Name",\
        "MailingAddr":  "Mailing Address",\
        "Phone":        "Phone",\
        "Neighborhood": "Neighborhood",\
    \}\
    raw_map = \{k: v for k, v in raw_map.items() if k and k in filtered.columns\}\
    raw_bldgs = filtered[list(raw_map)].copy()\
    raw_bldgs.columns = list(raw_map.values())\
\
    # \uc0\u9472 \u9472  VIOLATIONS (Phase 2) \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \
    print("\\n=== VIOLATIONS (Phase 2) ===")\
    viol_integrated = False\
    viol_by_bbl = pd.DataFrame(columns=["bbl_norm", "ViolCount", "PestFlag"])\
\
    if viols is not None:\
        vbbl = col(viols, "bbl")\
        vdesc= col(viols, "novdescription", "violationdescription", "description")\
        vstat= col(viols, "currentstatus", "status", "violationstatus")\
\
        if vbbl:\
            viols["bbl_norm"] = viols[vbbl].apply(normalize_bbl)\
            pipe_bbls = set(filtered["bbl_norm"].dropna())\
            viols = viols[viols["bbl_norm"].isin(pipe_bbls)].copy()\
            log("Violations matched to pipeline BBLs", len(viols))\
\
            if vstat:\
                open_mask = viols[vstat].str.strip().str.upper().isin(\
                    \{"OPEN", "ACTIVE", "VIOLATION ISSUED", "V*"\})\
                viols = viols[open_mask]\
                log("Violations (Open only)", len(viols))\
\
            pest_re = r"RODENT|ROACH|VERMIN|PEST|BED.?BUG|MOUSE|RAT|INFESTATION"\
            if vdesc:\
                viols["PestFlag"] = viols[vdesc].str.upper().str.contains(\
                    pest_re, na=False, regex=True).astype(int)\
            else:\
                viols["PestFlag"] = 0\
\
            viol_by_bbl = viols.groupby("bbl_norm").agg(\
                ViolCount=("bbl_norm", "count"),\
                PestFlag =("PestFlag", "max")\
            ).reset_index()\
            viol_integrated = True\
            log("Violations aggregated by BBL", len(viol_by_bbl))\
        else:\
            log("Violations", "SKIPPED", "No BBL column found in violations file")\
    else:\
        log("Violations", "SKIPPED", "hpd_violations.csv not provided")\
\
    # \uc0\u9472 \u9472  PHASE 2 SCORING \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \
    print("\\n=== PHASE 2 SCORING ===")\
    scored = filtered.copy()\
    if viol_integrated:\
        scored = scored.merge(viol_by_bbl, on="bbl_norm", how="left")\
        scored["ViolCount"] = scored["ViolCount"].fillna(0)\
        scored["PestFlag"]  = scored["PestFlag"].fillna(0)\
    else:\
        scored["ViolCount"] = np.nan\
        scored["PestFlag"]  = np.nan\
\
    # Retail flag\
    if p_lu and p_lu in scored.columns:\
        scored["RetailFlag"] = (scored[p_lu].astype(str).str.strip() == "04").astype(int)\
    elif p_retail and p_retail in scored.columns:\
        scored["RetailFlag"] = (pd.to_numeric(scored[p_retail], errors="coerce") > 0).astype(int)\
    else:\
        scored["RetailFlag"] = 0\
\
    # Building age\
    if p_yb and p_yb in scored.columns:\
        scored["BldgAge"] = 2026 - pd.to_numeric(scored[p_yb], errors="coerce")\
    else:\
        scored["BldgAge"] = np.nan\
\
    # Aggregate for scoring\
    grp2 = scored.groupby("_norm").agg(\
        AgentName    =("AgentName",   "first"),\
        N_Bldgs      =(p_units,       "count"),\
        TotalUnits   =(p_units,       "sum"),\
        AvgUnits     =(p_units,       "mean"),\
        TotalViols   =("ViolCount",   "sum"),\
        NRetail      =("RetailFlag",  "sum"),\
        Neighborhoods=("Neighborhood", join_uniq),\
        ZIPs         =(zip_col,       join_zips) if zip_col else ("AgentName", lambda _: ""),\
    ).reset_index()\
\
    # Re-join Phase 1 flags + contact info\
    grp2 = grp2.merge(\
        grp[["_norm", "PortfolioFlag", "MultiZIPFlag", "DensityFlag",\
             "Mailing", "Phone"]],\
        on="_norm", how="left"\
    )\
\
    # Violation score \'97 top quartile per neighborhood\
    if viol_integrated:\
        grp2["_hood1"] = grp2["Neighborhoods"].str.split(" | ").str[0]\
        q75_map = grp2.groupby("_hood1")["TotalViols"].transform(\
            lambda x: x.quantile(0.75))\
        grp2["ViolScore"] = (grp2["TotalViols"] > q75_map).astype(int) * 25\
        vscore_note = "top-quartile per neighborhood (25 pts)"\
    else:\
        grp2["ViolScore"] = 0\
        vscore_note = "0 (violations file absent)"\
\
    # Portfolio strength (30 pts)\
    grp2["PortScore"] = (grp2["N_Bldgs"] >= 5).astype(int) * 15 + \\\
                        (grp2["TotalUnits"] >= 300).astype(int) * 15\
\
    # Revenue potential (25 pts)\
    grp2["RevScore"] = grp2["AvgUnits"].between(40, 100).astype(int) * 15 + \\\
                       (grp2["NRetail"] > 0).astype(int) * 10\
\
    # Route efficiency (20 pts)\
    grp2["DensScore"] = (grp2["DensityFlag"] == "Yes").astype(int) * 20\
\
    # Total capped at 100\
    grp2["TotalScore"] = (grp2["PortScore"] + grp2["ViolScore"] +\
                          grp2["RevScore"] + grp2["DensScore"]).clip(upper=100)\
    grp2["Tier"] = grp2["TotalScore"].apply(\
        lambda s: "A" if s >= 75 else ("B" if s >= 50 else "C"))\
\
    score_notes = (\
        f"Join: BBL primary (\{n_bbl\} rows) + address+ZIP fallback (\{n_fallback\} rows). "\
        f"Violation score: \{vscore_note\}. "\
        f"Retail: PLUTO LandUse=04 or RetailArea>0. "\
        f"Density: same-ZIP heuristic."\
    )\
\
    ranked_p2 = pd.DataFrame(\{\
        "Managing Agent Name":          grp2["AgentName"],\
        "Mailing Address":              grp2["Mailing"],\
        "Phone":                        grp2["Phone"],\
        "Neighborhood":                 grp2["Neighborhoods"],\
        "ZIP Code(s)":                  grp2["ZIPs"],\
        "# Buildings Managed":          grp2["N_Bldgs"],\
        "Total Units":                  grp2["TotalUnits"].astype(int),\
        "Avg Units per Building":       grp2["AvgUnits"].round(1),\
        "Portfolio Operator Flag (5+)": grp2["PortfolioFlag"],\
        "Multi-ZIP Flag":               grp2["MultiZIPFlag"],\
        "Density Flag":                 grp2["DensityFlag"],\
        "Violation Score":              grp2["ViolScore"].astype(int),\
        "Revenue Potential Score":      grp2["RevScore"].astype(int),\
        "Density Score":                grp2["DensScore"].astype(int),\
        "Total Priority Score (0-100)": grp2["TotalScore"].astype(int),\
        "Recommended Tier (A/B/C)":     grp2["Tier"],\
        "Notes":                        score_notes,\
    \}).drop_duplicates(subset=["Managing Agent Name"])\
\
    ranked_p2 = ranked_p2.sort_values(\
        ["Total Priority Score (0-100)", "# Buildings Managed"],\
        ascending=[False, False]\
    ).reset_index(drop=True)\
    log("Phase 2 ranked managers", len(ranked_p2))\
\
    # Raw buildings scored\
    raw2_map = \{\
        "BuildingAddr": "Address",\
        "bbl_norm":     "BBL (if available)",\
        boro_col:       "Borough",\
        zip_col:        "ZIP",\
        p_units:        "UnitsRes",\
        "AgentName":    "Managing Agent Name",\
        "Neighborhood": "Neighborhood",\
        "ViolCount":    "Violation Count (if available)",\
        "PestFlag":     "Rodent/Pest Signal (if available)",\
        "RetailFlag":   "Ground-floor Retail Flag (if inferred)",\
        "BldgAge":      "Building Age (if available)",\
    \}\
    raw2_map = \{k: v for k, v in raw2_map.items() if k and k in scored.columns\}\
    raw_bldgs2 = scored[list(raw2_map)].copy()\
    raw_bldgs2.columns = list(raw2_map.values())\
\
    # \uc0\u9472 \u9472  DATA_NOTES rows \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \
    cs_contacts = sha256(os.path.join(data_dir, "hpd_contacts.csv"))\
    cs_regs     = sha256(os.path.join(data_dir, "hpd_registrations.csv"))\
    cs_pluto    = sha256(os.path.join(data_dir, "pluto.csv"))\
    cs_viols    = sha256(os.path.join(data_dir, "hpd_violations.csv")) \\\
                  if Path(os.path.join(data_dir, "hpd_violations.csv")).exists() else "absent"\
\
    dn_rows = [\
        ["Field", "Value"],\
        ["Run timestamp",         run_start],\
        ["HPD Contacts source",   SOQL_URLS["hpd_contacts.csv"]],\
        ["HPD Contacts SHA-256",  cs_contacts],\
        ["HPD Regs source",       SOQL_URLS["hpd_registrations.csv"]],\
        ["HPD Regs SHA-256",      cs_regs],\
        ["PLUTO source",          SOQL_URLS["pluto.csv"]],\
        ["PLUTO SHA-256",         cs_pluto],\
        ["HPD Violations source", SOQL_URLS["hpd_violations.csv"]],\
        ["HPD Violations SHA-256",cs_viols],\
        ["Target ZIPs",           ", ".join(sorted(TARGET_ZIPS))],\
        ["UnitsRes filter",       f"\{MIN_UNITS\}\'96\{MAX_UNITS\}"],\
        ["Primary join key",      "BBL (10-digit: BoroID[1]+Block[5]+Lot[4])"],\
        ["Fallback join key",     "Normalized address + ZIP"],\
        ["BBL-join rows",         str(n_bbl)],\
        ["Address+ZIP fallback",  str(n_fallback)],\
        ["Density method",        "\uc0\u8805 3 buildings in same ZIP per agent"],\
        ["Retail flag source",    "PLUTO LandUse=04 or RetailArea>0"],\
        ["Violations integrated", str(viol_integrated)],\
        ["Violation score method","Top quartile (75th pct) open violations / neighborhood = 25 pts"],\
    ]\
\
    # \uc0\u9472 \u9472  WRITE EXCEL \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \
    print("\\n=== WRITE EXCEL ===")\
    run_ts = datetime.now().strftime("%Y-%m-%d %H:%M")\
    note_p1 = (\
        f"LIVE DATA | \{run_ts\} | "\
        f"HPD Contacts feu5-w2e2 \'d7 HPD Regs tesw-yqqr \'d7 PLUTO 64uk-42ks | "\
        f"BBL join (addr+ZIP fallback) | No synthetic data."\
    )\
    note_p2 = note_p1 + f" | Violations: \{'integrated' if viol_integrated else 'absent (score=0)'\}."\
\
    def write_workbook(path: str, sheets: list[tuple]) -> str:\
        wb = Workbook()\
        first = True\
        for title, df, color, note_txt in sheets:\
            ws = wb.active if first else wb.create_sheet(title)\
            if first:\
                ws.title = title\
                first = False\
            _write_sheet(ws, df, hdr_color=color, note=note_txt)\
        wb.save(path)\
        ck = sha256(path)\
        log(f"Wrote \{Path(path).name\}", f"\{Path(path).stat().st_size // 1024\} KB",\
            f"sha256=\{ck[:16]\}\'85")\
        return ck\
\
    # Data_Notes + Run_Log DataFrames\
    dn_df  = pd.DataFrame(dn_rows[1:], columns=dn_rows[0])\
    rl_df  = pd.DataFrame(_run_log)   # current log snapshot\
\
    p1_path = os.path.join(output_dir, "NYC_Multifamily_TargetManagers_Phase1.xlsx")\
    ck_p1 = write_workbook(p1_path, [\
        ("Ranked_Managers",   ranked_p1,  "1F4E79", note_p1),\
        ("Raw_Buildings",     raw_bldgs,  "14532D", note_p1),\
        ("Data_Notes",        dn_df,      "1F4E79", ""),\
        ("Run_Log",           rl_df,      "5D4037", ""),\
    ])\
\
    # Refresh run log for Phase 2 workbook (includes Phase 1 write event)\
    rl_df2 = pd.DataFrame(_run_log)\
    p2_path = os.path.join(output_dir, "NYC_Multifamily_TargetManagers_Phase2_Scored.xlsx")\
    ck_p2 = write_workbook(p2_path, [\
        ("Ranked_Managers_Scored",  ranked_p2,   "1F4E79", note_p2),\
        ("Raw_Buildings_Scored",    raw_bldgs2,  "7B1FA2", note_p2),\
        ("Data_Notes",              dn_df,       "1F4E79", ""),\
        ("Run_Log",                 rl_df2,      "5D4037", ""),\
    ])\
\
    # Tier color\
    from openpyxl import load_workbook as _lw\
    for path in [p2_path]:\
        wb2 = _lw(path)\
        ws_r = wb2["Ranked_Managers_Scored"]\
        tier_ci = list(ranked_p2.columns).index("Recommended Tier (A/B/C)") + 1\
        for ri in range(3, len(ranked_p2) + 3):\
            c = ws_r.cell(ri, tier_ci)\
            if c.value in TIER_CLR:\
                c.fill = PatternFill("solid", fgColor=TIER_CLR[c.value])\
                c.font = Font(bold=True)\
        wb2.save(path)\
\
    # \uc0\u9472 \u9472  FINAL REPORT \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \
    print("\\n" + "=" * 70)\
    print("COMPLETE")\
    print("=" * 70)\
    print(f"  Phase 1  : \{p1_path\}")\
    print(f"             SHA-256: \{ck_p1\}")\
    print(f"  Phase 2  : \{p2_path\}")\
    print(f"             SHA-256: \{ck_p2\}")\
\
    print("\\nNeighborhood summary:")\
    for hood in sorted(set(ZIP_TO_NEIGHBORHOOD.values())):\
        nb = filtered[filtered["Neighborhood"] == hood].shape[0]\
        nm = ranked_p1[ranked_p1["Neighborhood"].str.contains(\
            hood.split("(")[0].strip(), na=False, regex=False)].shape[0]\
        print(f"  \{hood:<45\} buildings=\{nb:>4\}  managers=\{nm:>3\}")\
\
    print("\\nPhase 2 tier breakdown:")\
    for tier, cnt in ranked_p2["Recommended Tier (A/B/C)"].value_counts().sort_index().items():\
        print(f"  Tier \{tier\}: \{cnt\}")\
\
    print("\\nRun log (all steps):")\
    for e in _run_log:\
        print(f"  [\{e['Timestamp']\}] \{e['Step']:<52\} \{e['Value']\}")\
\
\
# \uc0\u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \
# ENTRY POINT\
# \uc0\u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \u9472 \
\
def main():\
    ap = argparse.ArgumentParser(\
        description="NYC Multifamily Pipeline \'97 local CSV only, no network calls.",\
        formatter_class=argparse.RawDescriptionHelpFormatter,\
        epilog=__doc__,\
    )\
    ap.add_argument("--data-dir",   default=".",\
                    help="Directory containing pre-downloaded CSVs")\
    ap.add_argument("--output-dir", default=".",\
                    help="Directory for output Excel files")\
    args = ap.parse_args()\
    run(args.data_dir, args.output_dir)\
\
\
if __name__ == "__main__":\
    main()\
}